"""
Void System — Excel Client Seed Script

Normalizes and imports 478 clients from the CRM Excel export into the Neon PostgreSQL database.
Handles: city normalization, CPF cleaning, phone formatting, lead source mapping, UUID generation.

Usage:
  DATABASE_URL="postgresql://..." python3 scripts/seed-clients.py
"""

import os
import sys
import uuid
import json
from datetime import datetime

try:
    import openpyxl
    import psycopg2
except ImportError:
    print("Installing required packages...")
    os.system("pip3 install openpyxl psycopg2-binary")
    import openpyxl
    import psycopg2


EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'void system v1.0', 'loveable', 'novo_relatrio_04-02-2026.xlsx')
DATABASE_URL = os.environ.get('DATABASE_URL') or os.environ.get('NETLIFY_DATABASE_URL')

if not DATABASE_URL:
    print("ERROR: DATABASE_URL environment variable is required.")
    sys.exit(1)


# Lead source mapping from Portuguese CRM to our enum
LEAD_SOURCE_MAP = {
    'Instagram': 'instagram',
    'Facebook': 'facebook',
    'Google': 'google',
    'Indicação': 'indicacao',
    'Evento': 'evento',
    'Passou na frente da Void': 'passante',
    'outro': 'outro',
}

# City normalization map for known duplicates
CITY_NORMALIZE = {
    'curitiba': 'Curitiba',
    'Curitiba ': 'Curitiba',
    'Campo largo ': 'Campo Largo',
    'Pinhais ': 'Pinhais',
    'Joinville ': 'Joinville',
    'São Jose dos Pinhais': 'São José dos Pinhais',
    'Sao Jose dos Pinhais': 'São José dos Pinhais',
}


def normalize_city(city):
    """Normalize city names: trim, fix casing, resolve known duplicates."""
    if not city:
        return None

    city = str(city).strip()

    if city in CITY_NORMALIZE:
        return CITY_NORMALIZE[city]

    # Title case for unknown cities
    return ' '.join(
        word.capitalize() if len(word) > 2 else word.lower()
        for word in city.split()
    )


def clean_cpf(cpf):
    """Remove all non-digit characters from CPF."""
    if not cpf:
        return None
    return ''.join(c for c in str(cpf) if c.isdigit()) or None


def clean_phone(phone):
    """Normalize phone to digits with country code."""
    if not phone:
        return None
    digits = ''.join(c for c in str(phone) if c.isdigit())
    if digits.startswith('55') and len(digits) >= 12:
        return f'+{digits}'
    if len(digits) >= 10:
        return f'+55{digits}'
    return str(phone).strip()


def parse_date(val):
    """Parse birth date to ISO string."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    try:
        return str(val).strip()
    except:
        return None


def ts_to_iso(val):
    """Convert datetime to ISO string for timestamps."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    return str(val)


def main():
    print(f"📂 Loading Excel from: {os.path.abspath(EXCEL_PATH)}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Report']

    print(f"📊 Found {ws.max_row - 1} client records")

    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()

    # Create tables if they don't exist (Drizzle should handle this, but just in case)
    inserted = 0
    skipped = 0
    errors = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        try:
            cpf_key, cpf_raw, nome, nascimento, telefone, email = row[0:6]
            mora_cwb, cep, bairro, cidade, profissao, origem_lead = row[6:12]
            criado_por, codigo, titulo, status, criador = row[12:17]
            finalizado_em, criado_em, atualizado_em, expirado_em, vencido = row[17:22]

            if not email:
                skipped += 1
                continue

            email_clean = str(email).strip().lower()
            cpf_clean = clean_cpf(cpf_raw)
            phone_clean = clean_phone(telefone)
            city_normalized = normalize_city(cidade)
            lead_source = LEAD_SOURCE_MAP.get(str(origem_lead).strip(), 'outro') if origem_lead else None
            birth_date = parse_date(nascimento)
            crm_created = ts_to_iso(criado_em)
            crm_updated = ts_to_iso(atualizado_em)

            client_id = str(uuid.uuid4())

            cur.execute("""
                INSERT INTO clients (
                    id, email, full_name, phone, cpf, role,
                    birth_date, address_neighborhood, address_city,
                    address_cep, address_city_normalized,
                    profession, lead_source,
                    imported_from_crm, crm_code, crm_created_at, crm_updated_at,
                    xp, level, total_spent, total_sessions,
                    sessions_float, sessions_massage, sessions_combo,
                    life_cycle_stage, created_at, updated_at
                ) VALUES (
                    %s, %s, %s, %s, %s, 'client',
                    %s, %s, %s,
                    %s, %s,
                    %s, %s,
                    TRUE, %s, %s, %s,
                    0, 'iniciado', 0, 0,
                    0, 0, 0,
                    'imported', NOW(), NOW()
                )
                ON CONFLICT (email) DO NOTHING
            """, (
                client_id,
                email_clean,
                str(nome).strip() if nome else 'Unknown',
                phone_clean,
                cpf_clean,
                birth_date,
                str(bairro).strip() if bairro else None,
                str(cidade).strip() if cidade else None,
                str(cep).strip() if cep else None,
                city_normalized,
                str(profissao).strip() if profissao else None,
                lead_source,
                str(codigo) if codigo else None,
                crm_created,
                crm_updated,
            ))

            if cur.rowcount > 0:
                inserted += 1

                # Auto-tag based on lead source
                tags_to_add = []
                if lead_source:
                    tags_to_add.append(f'origem:{lead_source}')
                if city_normalized:
                    tags_to_add.append(f'cidade:{city_normalized.lower()}')
                if profissao:
                    tags_to_add.append(f'profissao:{str(profissao).strip().lower()}')

                for tag in tags_to_add:
                    tag_id = str(uuid.uuid4())
                    cur.execute("""
                        INSERT INTO client_tags (id, client_id, tag, created_at)
                        VALUES (%s, %s, %s, NOW())
                    """, (tag_id, client_id, tag))
            else:
                skipped += 1

        except Exception as e:
            errors += 1
            print(f"  ❌ Row {row_idx}: {e}")

    conn.commit()
    cur.close()
    conn.close()

    print(f"\n✅ Import complete!")
    print(f"   Inserted: {inserted}")
    print(f"   Skipped (duplicate/empty): {skipped}")
    print(f"   Errors: {errors}")


if __name__ == '__main__':
    main()
